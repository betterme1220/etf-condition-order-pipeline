from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import uuid4
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from config import get_portfolio_source


class PortfolioOrderBuilderError(Exception):
    """从 portfolio-monitor.xlsx 构造内部订单失败。"""

    pass


@dataclass(frozen=True)
class PortfolioRow:
    row_number: int
    values: dict[str, Any]
    number_formats: dict[str, str] = field(default_factory=dict)


DEFAULT_COLUMN_MAPPING = {
    "security_code": "代码",
    "security_name": "名称",
    "asset_type": "类型",
    "asset_class": "资产大类",
    "current_position_qty": "持仓数量",
    "cost_price": "成本价",
    "latest_price": "最新价格/净值",
    "valuation_date": "估值日期/到期",
    "market_value": "持仓市值",
    "position_weight_pct": "仓位比例",
    "buy_channel": "买入渠道",
    "sold_or_redeemed": "是否已赎回/已卖出",
    "status_mark": "状态标记",
    "target_weight_pct": "目标仓位%",
    "buy_date": "买入日期",
    "profit_pct": "浮动盈亏%",
    "drawdown_pct": "单只回撤%",
    "trade_status": "交易状态",
    "special_mark": "特殊标记",
    "industry": "行业",
    "project_total_capital": "项目总资金",
    "batch_buy_pct": "单次买入比例%",
    "batch_sell_pct": "单次卖出比例%",
    "weekly_max_buy_batches": "每周最大买入次数",
    "weekly_used_buy_batches": "本周已买入次数",
    "allow_sell_to_zero": "允许卖到0",
    "lot_size": "最小交易单位",
    "tick_size": "价格最小变动",
}


def _get_portfolio_cfg() -> dict[str, Any]:
    """兼容新版 portfolio_source.path 和旧版 source.path。"""

    cfg = get_portfolio_source()

    if isinstance(cfg.get("portfolio_source"), dict):
        return cfg["portfolio_source"]

    if isinstance(cfg.get("source"), dict):
        normalized = dict(cfg)
        source_cfg = cfg["source"]

        if "path" not in normalized and source_cfg.get("path"):
            normalized["path"] = source_cfg["path"]

        if "sheet" not in normalized and source_cfg.get("sheet"):
            normalized["sheet"] = source_cfg["sheet"]

        if "sheet_name" not in normalized and source_cfg.get("sheet_name"):
            normalized["sheet_name"] = source_cfg["sheet_name"]

        if "workbook_path" not in normalized and source_cfg.get("workbook_path"):
            normalized["workbook_path"] = source_cfg["workbook_path"]

        if "file_path" not in normalized and source_cfg.get("file_path"):
            normalized["file_path"] = source_cfg["file_path"]

        return normalized

    return cfg


def _get_column_mapping() -> dict[str, str]:
    """读取列映射。

    兼容两套配置：

    新字段名：
      security_code, security_name, asset_type, buy_channel,
      current_position_qty, latest_price ...

    旧字段名：
      code, name, type, channel, quantity, current_price ...

    注意：
    旧字段名只用于补齐新字段名，不再原样加入 mapping，
    避免同一 Excel 列被 channel 覆盖 buy_channel。
    """

    cfg = _get_portfolio_cfg()
    mapping = dict(DEFAULT_COLUMN_MAPPING)

    cfg_mapping = cfg.get("column_mapping", {})
    if not isinstance(cfg_mapping, dict):
        return mapping

    legacy_aliases = {
        "code": "security_code",
        "name": "security_name",
        "type": "asset_type",
        "channel": "buy_channel",
        "status": "status_mark",
        "quantity": "current_position_qty",
        "current_price": "latest_price",
        "current_weight": "position_weight_pct",
        "target_weight": "target_weight_pct",
    }

    for raw_key, header_name in cfg_mapping.items():
        logical_key = legacy_aliases.get(raw_key, raw_key)

        if logical_key in DEFAULT_COLUMN_MAPPING:
            mapping[logical_key] = header_name

    return mapping


def _get_workbook_path(portfolio_path: str | Path | None = None) -> Path:
    if portfolio_path is not None:
        return Path(portfolio_path)

    cfg = _get_portfolio_cfg()
    raw_path = cfg.get("path") or cfg.get("workbook_path") or cfg.get("file_path")

    if not raw_path:
        raise PortfolioOrderBuilderError(
            "portfolio_source.yaml 未配置 portfolio_source.path，"
            "且调用时未传 portfolio_path。"
        )

    return Path(str(raw_path))


def _get_sheet_name(sheet_name: str | None = None) -> str | None:
    if sheet_name:
        return sheet_name

    cfg = _get_portfolio_cfg()
    return cfg.get("sheet") or cfg.get("sheet_name")


def _get_max_header_scan_rows(max_header_scan_rows: int | None = None) -> int:
    if max_header_scan_rows is not None:
        if max_header_scan_rows <= 0:
            raise PortfolioOrderBuilderError(
                f"max_header_scan_rows 必须大于0: {max_header_scan_rows}"
            )
        return max_header_scan_rows

    cfg = _get_portfolio_cfg()
    raw_value = cfg.get("max_header_scan_rows", 50)

    try:
        value = int(raw_value)
    except Exception as exc:
        raise PortfolioOrderBuilderError(
            f"portfolio_source.max_header_scan_rows 不是有效整数: {raw_value!r}"
        ) from exc

    if value <= 0:
        raise PortfolioOrderBuilderError(
            f"portfolio_source.max_header_scan_rows 必须大于0: {raw_value!r}"
        )

    return value


def _get_percentage_input_mode(percentage_input_mode: str | None = None) -> str:
    """读取百分比解释模式。

    支持：
    - auto:
      如果 Excel 单元格格式含 %，则 0.1 解释为 10%；
      如果是普通数字 0.1，则解释为 0.1%。

    - percentage_points:
      始终把数字本身解释为百分数点位。
      例如 10 = 10%，0.1 = 0.1%。

    - excel_fraction:
      始终把数字当作 Excel 百分比小数。
      例如 0.1 = 10%。
    """

    if percentage_input_mode:
        mode = percentage_input_mode
    else:
        cfg = _get_portfolio_cfg()
        mode = str(cfg.get("percentage_input_mode", "auto"))

    allowed = {"auto", "percentage_points", "excel_fraction"}
    if mode not in allowed:
        raise PortfolioOrderBuilderError(
            f"percentage_input_mode 只能是 {sorted(allowed)}，当前: {mode!r}"
        )

    return mode



def _normalize_header(value: Any) -> str:
    """用于读取 Excel 表头的基础规范化。"""

    return str(value).strip() if value is not None else ""

def _normalize_header_key(value: Any) -> str:
    """用于表头匹配：忽略空格、换行、制表符、全角空格。"""

    text = _normalize_header(value)
    return (
        text.replace(" ", "")
        .replace("\t", "")
        .replace("\r", "")
        .replace("\n", "")
        .replace("\u3000", "")
    )


def _normalize_text(value: Any) -> str | None:
    if value is None:
        return None

    text = str(value).strip()
    return text or None


def _normalize_compact_text(value: Any) -> str | None:
    text = _normalize_text(value)
    if text is None:
        return None
    return text.replace(" ", "").replace("\u3000", "")


def _normalize_upper_text(value: Any) -> str | None:
    text = _normalize_compact_text(value)
    if text is None:
        return None
    return text.upper()


def _normalize_security_code(value: Any) -> str:
    if value is None or value == "":
        raise PortfolioOrderBuilderError("证券代码为空")

    if isinstance(value, float):
        text = str(int(value))
    else:
        text = str(value).strip()

    if "." in text:
        left, right = text.split(".", 1)
        if right == "0":
            text = left

    text = text.strip()

    if not text:
        raise PortfolioOrderBuilderError("证券代码为空")

    return text.zfill(6)


def _to_decimal(value: Any, field_name: str) -> Decimal:
    if value is None or value == "":
        raise PortfolioOrderBuilderError(f"字段为空: {field_name}")

    text = str(value).strip().replace(",", "")

    if text.endswith("%"):
        text = text[:-1].strip()

    try:
        return Decimal(text)
    except Exception as exc:
        raise PortfolioOrderBuilderError(
            f"字段不是有效数字: {field_name}={value!r}"
        ) from exc


def _to_positive_decimal(value: Any, field_name: str) -> Decimal:
    result = _to_decimal(value, field_name)
    if result <= 0:
        raise PortfolioOrderBuilderError(f"字段必须大于0: {field_name}={value!r}")
    return result


def _to_int(value: Any, field_name: str) -> int:
    if value is None or value == "":
        raise PortfolioOrderBuilderError(f"字段为空: {field_name}")

    try:
        return int(Decimal(str(value).strip().replace(",", "")))
    except Exception as exc:
        raise PortfolioOrderBuilderError(
            f"字段不是有效整数: {field_name}={value!r}"
        ) from exc


def _to_non_negative_int(value: Any, field_name: str) -> int:
    result = _to_int(value, field_name)
    if result < 0:
        raise PortfolioOrderBuilderError(f"字段不能为负数: {field_name}={value!r}")
    return result


def _to_positive_int(value: Any, field_name: str) -> int:
    result = _to_int(value, field_name)
    if result <= 0:
        raise PortfolioOrderBuilderError(f"字段必须大于0: {field_name}={value!r}")
    return result


def _has_percent_number_format(number_format: str | None) -> bool:
    return "%" in str(number_format or "")


def _to_pct_decimal(
    value: Any,
    field_name: str,
    *,
    number_format: str | None = None,
    percentage_input_mode: str = "auto",
) -> Decimal:
    """把 Excel 百分比字段转换为百分数点位。

    返回示例：
    - 10      -> Decimal("10")
    - "10%"   -> Decimal("10")
    - 0.1 且单元格格式为百分比 -> Decimal("10")
    - 0.1 且单元格格式不是百分比 -> Decimal("0.1")
    """

    if value is None or value == "":
        raise PortfolioOrderBuilderError(f"字段为空: {field_name}")

    raw_text = str(value).strip()
    explicit_percent_text = raw_text.endswith("%")

    raw = _to_positive_decimal(value, field_name)

    if explicit_percent_text:
        return raw

    if percentage_input_mode == "excel_fraction":
        return raw * Decimal("100")

    if percentage_input_mode == "percentage_points":
        return raw

    if percentage_input_mode == "auto":
        if _has_percent_number_format(number_format):
            return raw * Decimal("100")
        return raw

    raise PortfolioOrderBuilderError(
        f"未知 percentage_input_mode: {percentage_input_mode!r}"
    )


def _format_decimal(value: Decimal) -> str:
    return format(value.normalize(), "f")


def _to_pct_value(
    value: Any,
    field_name: str,
    *,
    number_format: str | None = None,
    percentage_input_mode: str = "auto",
) -> str:
    return _format_decimal(
        _to_pct_decimal(
            value,
            field_name,
            number_format=number_format,
            percentage_input_mode=percentage_input_mode,
        )
    )


def _to_bool(value: Any, field_name: str) -> bool:
    if isinstance(value, bool):
        return value

    if value is None or value == "":
        raise PortfolioOrderBuilderError(f"字段为空: {field_name}")

    text = str(value).strip().lower()

    true_values = {"是", "yes", "y", "true", "1", "允许", "可", "可以"}
    false_values = {"否", "no", "n", "false", "0", "不允许", "不可", "不可以"}

    if text in true_values:
        return True

    if text in false_values:
        return False

    raise PortfolioOrderBuilderError(
        f"字段不是有效布尔值: {field_name}={value!r}，请填 是/否。"
    )


def _is_sold_or_redeemed(value: Any) -> bool:
    """判断是否已赎回/已卖出。

    允许 Excel 中填写：
    - 是 / true / 1 / 已赎回 / 已卖出 / 赎回 / 卖出 / 已清仓 / 清仓
    - 否 / false / 0 / 未赎回 / 未卖出 / 持有中 / 空值
    """

    if value is None or value == "":
        return False

    if isinstance(value, bool):
        return value

    text = str(value).strip().lower()

    true_values = {
        "是",
        "yes",
        "y",
        "true",
        "1",
        "已赎回",
        "已卖出",
        "赎回",
        "卖出",
        "已清仓",
        "清仓",
    }

    false_values = {
        "否",
        "no",
        "n",
        "false",
        "0",
        "未赎回",
        "未卖出",
        "未清仓",
        "持有",
        "持有中",
        "正常",
    }

    if text in true_values:
        return True

    if text in false_values:
        return False

    raise PortfolioOrderBuilderError(
        f"是否已赎回/已卖出 字段不是有效值: {value!r}，"
        "请填写 是/否、已赎回、已卖出、持有中。"
    )


def _is_disabled_trade_status(value: Any) -> bool:
    if value is None or value == "":
        return False

    text = str(value).strip()

    disabled_status = {
        "暂停",
        "停止",
        "禁用",
        "清仓观察",
        "不可交易",
        "暂停交易",
        "禁止交易",
    }

    return text in disabled_status


def _load_workbook_safely(path: Path):
    try:
        return load_workbook(path, data_only=True, read_only=True)
    except (InvalidFileException, BadZipFile, PermissionError, OSError, ValueError) as exc:
        raise PortfolioOrderBuilderError(
            f"无法读取持仓工作簿: {path}。"
            f"可能原因：文件损坏、格式不支持、被占用、加密或权限不足。"
            f"原始错误: {type(exc).__name__}: {exc}"
        ) from exc
    except Exception as exc:
        raise PortfolioOrderBuilderError(
            f"读取持仓工作簿时发生未知错误: {path}。"
            f"原始错误: {type(exc).__name__}: {exc}"
        ) from exc


def _find_header_row(
    ws,
    required_headers: set[str],
    *,
    max_header_scan_rows: int,
) -> tuple[int, dict[str, int]]:
    """在前 max_header_scan_rows 行内寻找表头行。

    表头匹配会忽略空格、换行、制表符、全角空格。
    """

    scan_rows = min(ws.max_row, max_header_scan_rows)
    required_header_keys = {_normalize_header_key(header) for header in required_headers}

    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=scan_rows),
        start=1,
    ):
        header_map: dict[str, int] = {}

        for cell in row:
            header_key = _normalize_header_key(cell.value)
            if header_key:
                header_map[header_key] = cell.column

        if required_header_keys.issubset(set(header_map.keys())):
            return row_number, header_map

    raise PortfolioOrderBuilderError(
        f"未找到有效表头行。"
        f"已扫描前 {scan_rows} 行，必须至少包含: {sorted(required_headers)}。"
        "请确认表头为单行文本表头，不使用合并单元格。"
        "如表头在更靠后位置，请在 portfolio_source.yaml 中提高 max_header_scan_rows。"
    )


def _load_portfolio_rows(
    *,
    portfolio_path: str | Path | None = None,
    sheet_name: str | None = None,
    max_header_scan_rows: int | None = None,
) -> list[PortfolioRow]:
    path = _get_workbook_path(portfolio_path)

    if not path.exists():
        raise PortfolioOrderBuilderError(f"持仓工作簿不存在: {path}")

    workbook = _load_workbook_safely(path)

    try:
        target_sheet_name = _get_sheet_name(sheet_name)
        if target_sheet_name:
            if target_sheet_name not in workbook.sheetnames:
                raise PortfolioOrderBuilderError(
                    f"工作簿中不存在工作表: {target_sheet_name}，"
                    f"可用工作表: {workbook.sheetnames}"
                )
            ws = workbook[target_sheet_name]
        else:
            ws = workbook[workbook.sheetnames[0]]

        column_mapping = _get_column_mapping()

        required_headers = {
            column_mapping["security_code"],
            column_mapping["security_name"],
        }

        header_row_number, header_map = _find_header_row(
            ws,
            required_headers,
            max_header_scan_rows=_get_max_header_scan_rows(max_header_scan_rows),
        )

        logical_by_column: dict[int, str] = {}
        for logical_name, header_name in column_mapping.items():
            header_key = _normalize_header_key(header_name)
            column_index = header_map.get(header_key)
            if column_index is not None:
                logical_by_column[column_index] = logical_name

        rows: list[PortfolioRow] = []

        for row_number, row in enumerate(
            ws.iter_rows(min_row=header_row_number + 1, max_row=ws.max_row),
            start=header_row_number + 1,
        ):
            values: dict[str, Any] = {
                logical_name: None for logical_name in column_mapping
            }
            number_formats: dict[str, str] = {}

            for cell in row:
                logical_name = logical_by_column.get(cell.column)
                if logical_name is None:
                    continue

                values[logical_name] = cell.value
                number_formats[logical_name] = str(
                    getattr(cell, "number_format", "") or ""
                )

            if values.get("security_code") in (None, ""):
                continue

            rows.append(
                PortfolioRow(
                    row_number=row_number,
                    values=values,
                    number_formats=number_formats,
                )
            )

        return rows

    finally:
        workbook.close()


def _find_security_row(
    security_code: str,
    *,
    portfolio_path: str | Path | None = None,
    sheet_name: str | None = None,
    max_header_scan_rows: int | None = None,
) -> PortfolioRow:
    target_code = _normalize_security_code(security_code)

    rows = _load_portfolio_rows(
        portfolio_path=portfolio_path,
        sheet_name=sheet_name,
        max_header_scan_rows=max_header_scan_rows,
    )

    for row in rows:
        row_code = _normalize_security_code(row.values.get("security_code"))
        if row_code == target_code:
            return row

    raise PortfolioOrderBuilderError(f"未在持仓表中找到证券代码: {target_code}")


def _assert_framework_allowed_etf(row: PortfolioRow) -> None:
    values = row.values

    security_code = _normalize_security_code(values.get("security_code"))

    asset_type_raw = _normalize_text(values.get("asset_type"))
    asset_type = _normalize_upper_text(values.get("asset_type"))

    buy_channel_raw = _normalize_text(values.get("buy_channel"))
    buy_channel = _normalize_compact_text(values.get("buy_channel"))

    sold_or_redeemed_raw = values.get("sold_or_redeemed")
    trade_status = _normalize_text(values.get("trade_status"))

    if not asset_type or "ETF" not in asset_type:
        raise PortfolioOrderBuilderError(
            f"当前框架只允许场内ETF: code={security_code}, 类型={asset_type_raw!r}"
        )

    allowed_channels = {"场内", "场内ETF", "场内基金", "交易所", "二级市场"}
    if buy_channel not in allowed_channels:
        raise PortfolioOrderBuilderError(
            f"当前框架只允许场内ETF: code={security_code}, 买入渠道={buy_channel_raw!r}"
        )

    if _is_sold_or_redeemed(sold_or_redeemed_raw):
        raise PortfolioOrderBuilderError(
            f"该标的已赎回或已卖出，不应生成条件单: code={security_code}"
        )

    if _is_disabled_trade_status(trade_status):
        raise PortfolioOrderBuilderError(
            f"该标的交易状态不允许生成条件单: "
            f"code={security_code}, 交易状态={trade_status!r}"
        )


def _make_system_order_id(
    security_code: str,
    *,
    template_code: str = "GRID",
    suffix: str | None = None,
) -> str:
    if suffix:
        safe_suffix = "".join(
            char for char in str(suffix).strip().upper() if char.isalnum()
        )
        if not safe_suffix:
            safe_suffix = uuid4().hex[:6].upper()
    else:
        safe_suffix = uuid4().hex[:6].upper()

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")

    return f"ETF-{timestamp}-{security_code}-{template_code}-{safe_suffix}"


def build_grid_order_from_portfolio(
    security_code: str,
    *,
    portfolio_path: str | Path | None = None,
    sheet_name: str | None = None,
    grid_pct: str | Decimal = "1.0",
    price_range_pct: str | Decimal = "3.0",
    allow_grid_without_monitor_window: bool = True,
    output_mode: str = "DRAFT",
    quantity_price_basis: str = "latest_price_or_nav",
    max_header_scan_rows: int | None = None,
    percentage_input_mode: str | None = None,
    allow_asymmetric_batch_pct_record_only: bool = False,
    order_id_suffix: str | None = None,
) -> dict[str, Any]:
    """从 portfolio-monitor.xlsx 中读取一只 ETF，生成网格交易内部订单。

    生成结果可直接传给：

        ManualAppChecklistGenerator().generate(order)

    注意：
    - 本函数只生成内部订单 dict。
    - 不连接券商。
    - 不自动点击 APP。
    - 不自动下单。
    """

    row = _find_security_row(
        security_code,
        portfolio_path=portfolio_path,
        sheet_name=sheet_name,
        max_header_scan_rows=max_header_scan_rows,
    )

    _assert_framework_allowed_etf(row)

    values = row.values
    number_formats = row.number_formats
    pct_mode = _get_percentage_input_mode(percentage_input_mode)

    normalized_code = _normalize_security_code(values.get("security_code"))
    security_name = _normalize_text(values.get("security_name")) or normalized_code

    latest_price = _to_positive_decimal(
        values.get("latest_price"),
        "最新价格/净值",
    )

    current_position_qty = _to_non_negative_int(
        values.get("current_position_qty"),
        "持仓数量",
    )

    project_total_capital = _to_positive_decimal(
        values.get("project_total_capital"),
        "项目总资金",
    )

    batch_buy_pct_decimal = _to_pct_decimal(
        values.get("batch_buy_pct"),
        "单次买入比例%",
        number_format=number_formats.get("batch_buy_pct"),
        percentage_input_mode=pct_mode,
    )
    batch_buy_pct = _format_decimal(batch_buy_pct_decimal)

    batch_sell_pct_raw = values.get("batch_sell_pct")
    batch_sell_pct_decimal: Decimal | None = None
    batch_sell_pct: str | None = None

    if batch_sell_pct_raw not in (None, ""):
        batch_sell_pct_decimal = _to_pct_decimal(
            batch_sell_pct_raw,
            "单次卖出比例%",
            number_format=number_formats.get("batch_sell_pct"),
            percentage_input_mode=pct_mode,
        )
        batch_sell_pct = _format_decimal(batch_sell_pct_decimal)

        if (
            batch_sell_pct_decimal != batch_buy_pct_decimal
            and not allow_asymmetric_batch_pct_record_only
        ):
            raise PortfolioOrderBuilderError(
                "当前 MVP 网格清单只支持买卖同一委托数量，"
                "不支持单次买入比例% 与 单次卖出比例% 不一致。"
                f"当前: 单次买入比例%={batch_buy_pct_decimal}, "
                f"单次卖出比例%={batch_sell_pct_decimal}。"
                "如只是想记录卖出比例而不用于本次清单，"
                "请显式设置 allow_asymmetric_batch_pct_record_only=True。"
            )

    weekly_max_buy_batches = _to_positive_int(
        values.get("weekly_max_buy_batches"),
        "每周最大买入次数",
    )

    weekly_used_buy_batches = _to_non_negative_int(
        values.get("weekly_used_buy_batches"),
        "本周已买入次数",
    )

    allow_sell_to_zero = _to_bool(
        values.get("allow_sell_to_zero"),
        "允许卖到0",
    )

    lot_size_raw = values.get("lot_size")
    tick_size_raw = values.get("tick_size")

    lot_size = _to_positive_int(
        lot_size_raw if lot_size_raw not in (None, "") else 100,
        "最小交易单位",
    )

    tick_size = _to_positive_decimal(
        tick_size_raw if tick_size_raw not in (None, "") else "0.001",
        "价格最小变动",
    )

    order: dict[str, Any] = {
        "system_order_id": _make_system_order_id(
            normalized_code,
            template_code="GRID",
            suffix=order_id_suffix,
        ),
        "template_key": "grid_trading",
        "security_code": normalized_code,
        "security_name": security_name,

        "initial_base_price": str(latest_price),
        "grid_pct": str(Decimal(str(grid_pct)).normalize()),
        "price_range_pct": str(Decimal(str(price_range_pct)).normalize()),

        "lot_size": lot_size,
        "tick_size": str(tick_size),

        "current_position_qty": current_position_qty,

        "project_total_capital": str(project_total_capital),
        "batch_buy_pct": batch_buy_pct,
        "batch_sell_pct": batch_sell_pct,

        "quantity_reference_price": str(latest_price),
        "quantity_price_basis": quantity_price_basis,

        "allow_sell_to_zero": allow_sell_to_zero,

        "weekly_max_buy_batches": weekly_max_buy_batches,
        "weekly_used_buy_batches": weekly_used_buy_batches,

        "allow_grid_without_monitor_window": allow_grid_without_monitor_window,
        "output_mode": output_mode,

        "source": {
            "portfolio_row_number": row.row_number,
            "security_name": security_name,
            "asset_type": values.get("asset_type"),
            "buy_channel": values.get("buy_channel"),
            "trade_status": values.get("trade_status"),
            "percentage_input_mode": pct_mode,
            "batch_sell_pct_policy": (
                "same_as_buy_pct"
                if batch_sell_pct_decimal == batch_buy_pct_decimal
                else (
                    "record_only_not_used_by_mvp"
                    if batch_sell_pct_decimal is not None
                    else "not_provided"
                )
            ),
            "lot_size_source": "excel" if lot_size_raw not in (None, "") else "default_100",
            "tick_size_source": (
                "excel" if tick_size_raw not in (None, "") else "default_0.001"
            ),
        },
    }

    return order


def build_grid_checklist_text_from_portfolio(
    security_code: str,
    *,
    portfolio_path: str | Path | None = None,
    sheet_name: str | None = None,
    grid_pct: str | Decimal = "1.0",
    price_range_pct: str | Decimal = "3.0",
    allow_grid_without_monitor_window: bool = True,
    output_mode: str = "DRAFT",
    max_header_scan_rows: int | None = None,
    percentage_input_mode: str | None = None,
    allow_asymmetric_batch_pct_record_only: bool = False,
    order_id_suffix: str | None = None,
) -> str:
    """从持仓 Excel 直接生成网格交易 APP 人工填写清单文本。"""

    from core.broker_adapter import ManualAppChecklistGenerator, render_checklist_text

    order = build_grid_order_from_portfolio(
        security_code,
        portfolio_path=portfolio_path,
        sheet_name=sheet_name,
        grid_pct=grid_pct,
        price_range_pct=price_range_pct,
        allow_grid_without_monitor_window=allow_grid_without_monitor_window,
        output_mode=output_mode,
        max_header_scan_rows=max_header_scan_rows,
        percentage_input_mode=percentage_input_mode,
        allow_asymmetric_batch_pct_record_only=allow_asymmetric_batch_pct_record_only,
        order_id_suffix=order_id_suffix,
    )

    result = ManualAppChecklistGenerator().generate(order)
    return render_checklist_text(result)